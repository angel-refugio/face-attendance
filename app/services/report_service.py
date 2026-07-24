import csv
import io
import logging
from datetime import datetime

logger = logging.getLogger(__name__)


def export_attendance_csv(records):
    output = io.StringIO()
    writer = csv.writer(output)

    writer.writerow(
        ["Empleado", "Departamento", "Fecha", "Hora", "Confianza", "Estado", "Notas"]
    )

    for record in records:
        writer.writerow(
            [
                record.employee.name,
                record.employee.department or "",
                record.check_in_time.strftime("%Y-%m-%d"),
                record.check_in_time.strftime("%H:%M:%S"),
                f"{record.confidence_score:.1f}%" if record.confidence_score else "",
                "Puntual" if record.status == "on_time" else "Tarde",
                record.notes or "",
            ]
        )

    output.seek(0)
    return output.getvalue()


def export_attendance_excel(records):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        logger.error("openpyxl no instalado")
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = "Asistencia"

    headers = [
        "Empleado",
        "Departamento",
        "Fecha",
        "Hora",
        "Confianza",
        "Estado",
        "Notas",
    ]
    ws.append(headers)

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for record in records:
        ws.append(
            [
                record.employee.name,
                record.employee.department or "",
                record.check_in_time.strftime("%Y-%m-%d"),
                record.check_in_time.strftime("%H:%M:%S"),
                f"{record.confidence_score:.1f}%" if record.confidence_score else "",
                "Puntual" if record.status == "on_time" else "Tarde",
                record.notes or "",
            ]
        )

    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
